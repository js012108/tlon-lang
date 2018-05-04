from django.core.files.base import ContentFile
from django.conf import settings
from django.utils import timezone
from .models import MigracionBaseCertificada, UsuarioSisben
import openpyxl


def subir_archivo(archivo):
	carpeta = '/media/files/base_certificada/migraciones/'
	directorio_base = getattr(settings, "BASE_DIR")
	extension = archivo.name.split('.')[-1]
	fecha = timezone.now().strftime("%Y%m%d_%H%M%S")
	nombre_archivo = 'import_' + fecha + '.' + extension

	path = directorio_base + carpeta + nombre_archivo

	archivo_generado = open(path, 'wb+')
	contenido_archivo = ContentFile(archivo.read())

	try:
		for chunk in contenido_archivo.chunks():
			archivo_generado.write(chunk)
		archivo_generado.close()

		return (True, nombre_archivo)
	except Exception as err:
		return (False, err)


def migrar_archivo(migracion):
	directorio_base = getattr(settings, "BASE_DIR")
	directorio = '/media/files/base_certificada/migraciones/'
	nombre_archivo = migracion.file_url

	ruta = directorio_base + directorio + nombre_archivo

	workbook = openpyxl.load_workbook(ruta, guess_types=True)
	sheet = workbook.get_active_sheet()
	max_filas = sheet.max_row + 1
	
	for fila in range(2, max_filas):

		numero_registro = sheet.cell(row=fila, column=1).value  # A
		documento_identidad = sheet.cell(row=fila, column=77).value  # BY
		tipo_documento = sheet.cell(row=fila, column=76).value  # BX

		primer_apellido = sheet.cell(row=fila, column=70).value  # BR
		segundo_apellido = sheet.cell(row=fila, column=71).value  # BS
		primer_nombre = sheet.cell(row=fila, column=72).value  # BT
		segundo_nombre = sheet.cell(row=fila, column=73).value  # BU

		genero = sheet.cell(row=fila, column=74).value  # BV
		fecha_nacimiento = sheet.cell(row=fila, column=78).value  # BZ
		edad_corte = sheet.cell(row=fila, column=98).value  # CT
		estado_civil = sheet.cell(row=fila, column=80).value  # CB

		parentesco = sheet.cell(row=fila, column=79).value  # CA
		extranjero = sheet.cell(row=fila, column=75).value  # BW
		ficha = sheet.cell(row=fila, column=2).value  # B
		puntaje_sisben = sheet.cell(row=fila, column=118).value  # DN
		departamento = sheet.cell(row=fila, column=3).value  # C
		municipio = sheet.cell(row=fila, column=4).value  # D
		zona = sheet.cell(row=fila, column=5).value  # E
		sector = sheet.cell(row=fila, column=6).value  # F
		manzana = sheet.cell(row=fila, column=8).value  # H
		comuna = sheet.cell(row=fila, column=9).value  # I
		barrio = sheet.cell(row=fila, column=10).value  # J
		vereda = sheet.cell(row=fila, column=11).value  # K
		direccion = sheet.cell(row=fila, column=12).value  # L
		telefono = sheet.cell(row=fila, column=20).value  # T
		tipo_vivienda = sheet.cell(row=fila, column=14).value  # N
		estrato = sheet.cell(row=fila, column=23).value  # W

		nivel_educativo = sheet.cell(row=fila, column=91).value  # CM
		tipo_establecimiento = sheet.cell(row=fila, column=89).value  # CK
		actividad_economica = sheet.cell(row=fila, column=92).value  # CN
		activo = True
		afrodescendiente = False

		nuevo_registro = UsuarioSisben(numero_registro=numero_registro,
		                               documento_identidad=documento_identidad, tipo_documento=tipo_documento,
		                               primer_apellido=primer_apellido, segundo_apellido=segundo_apellido,
		                               primer_nombre=primer_nombre, segundo_nombre=segundo_nombre,
		                               genero=genero, fecha_nacimiento=fecha_nacimiento,
		                               edad_corte=edad_corte, estado_civil=estado_civil,
		                               parentesco=parentesco, extranjero=extranjero,
		                               ficha=ficha, puntaje_sisben=puntaje_sisben,
		                               departamento=departamento, municipio=municipio,
		                               zona=zona, sector=sector, manzana=manzana,
		                               comuna=comuna, barrio=barrio, vereda=vereda,
		                               direccion=direccion, telefono=telefono,
		                               tipo_vivienda=tipo_vivienda, estrato=estrato,
		                               nivel_educativo=nivel_educativo,
		                               tipo_establecimiento=tipo_establecimiento,
		                               actividad_economica=actividad_economica,
		                               activo=activo,
									   afrodescendiente = afrodescendiente
		                               )
		nuevo_registro.save()